# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, send_from_directory, redirect, make_response
import sqlite3
import os
import datetime
import json
import io
from flask import send_file

CARPETA_ACTUAL = os.path.dirname(os.path.abspath(__file__))
CARPETA_RAIZ = os.path.dirname(CARPETA_ACTUAL)
RUTA_BD = os.path.join(CARPETA_ACTUAL, 'sistema_pat.db')
UPLOAD_FOLDER = os.path.join(CARPETA_ACTUAL, 'uploads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

CUATS = {'ENE-ABR': [2,5,8], 'MAY-AGO': [3,6,9], 'SEP-DIC': [1,4,7]}

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# -- CORS manual --
@app.after_request
def cors(r):
    r.headers['Access-Control-Allow-Origin'] = '*'
    r.headers['Access-Control-Allow-Headers'] = 'Content-Type'
    r.headers['Access-Control-Allow-Methods'] = 'GET,POST,PUT,PATCH,DELETE,OPTIONS'
    return r

@app.before_request
def preflight():
    if request.method == 'OPTIONS':
        r = make_response()
        r.headers['Access-Control-Allow-Origin'] = '*'
        r.headers['Access-Control-Allow-Headers'] = 'Content-Type'
        r.headers['Access-Control-Allow-Methods'] = 'GET,POST,PUT,PATCH,DELETE,OPTIONS'
        return r, 200

def db():
    conn = sqlite3.connect(RUTA_BD)
    conn.row_factory = sqlite3.Row
    return conn

# ============ STATIC FILES ============
@app.route('/')
def index(): return redirect('/paginas/login.html')

@app.route('/paginas/<path:f>')
def pg(f): return send_from_directory(os.path.join(CARPETA_RAIZ, 'paginas'), f)

@app.route('/css/<path:f>')
def css(f): return send_from_directory(os.path.join(CARPETA_RAIZ, 'css'), f)

@app.route('/js/<path:f>')
def js(f): return send_from_directory(os.path.join(CARPETA_RAIZ, 'js'), f)

@app.route('/assets/<path:f>')
def assets(f): return send_from_directory(os.path.join(CARPETA_RAIZ, 'assets'), f)

# ============ AUTH ============
@app.route('/api/login', methods=['POST'])
def login():
    d = request.get_json() or {}
    email = (d.get('email') or '').strip().lower()
    pwd = (d.get('password') or '').strip()
    if not email or not pwd:
        return jsonify(success=False, message='Correo y contrasena requeridos'), 400
    try:
        cn = db()
        u = cn.execute("SELECT * FROM usuarios WHERE LOWER(email)=? AND password=? AND activo=1", (email, pwd)).fetchone()
        cn.close()
        if u:
            return jsonify(success=True, id=u['id'], nombre=u['nombre'], apellidos=u['apellidos'],
                           email=u['email'], role=u['rol'], departamento=u['departamento_o_carrera'],
                           grupo=u['grupo_asignado'], tutor_id=u['tutor_id'])
        return jsonify(success=False, message='Correo o contrasena incorrectos'), 401
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ PERIODOS ============
@app.route('/api/periodos/activo')
def periodo_activo():
    try:
        cn = db()
        p = cn.execute("SELECT * FROM periodos WHERE activo=1 LIMIT 1").fetchone()
        cn.close()
        if not p: return jsonify(success=False, message='Sin periodo activo'), 404
        d = dict(p)
        d['cuatrimestres'] = CUATS.get(d['tipo'], [])
        return jsonify(success=True, periodo=d)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ CARRERAS ============
@app.route('/api/carreras')
def carreras():
    try:
        cn = db()
        rows = cn.execute("SELECT id,nombre,abreviatura FROM carreras ORDER BY nombre").fetchall()
        cn.close()
        return jsonify(success=True, carreras=[dict(r) for r in rows])
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ USUARIOS ============
@app.route('/api/usuarios', methods=['GET'])
def list_users():
    try:
        cn = db()
        rows = cn.execute("SELECT id,nombre,apellidos,email,rol,departamento_o_carrera,grupo_asignado,tutor_id,activo FROM usuarios ORDER BY rol,nombre").fetchall()
        cn.close()
        return jsonify(success=True, usuarios=[dict(r) for r in rows])
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/usuarios', methods=['POST'])
def create_user():
    d = request.get_json() or {}
    nombre = (d.get('nombre') or '').strip()
    apellidos = (d.get('apellidos') or '').strip()
    email = (d.get('email') or '').strip().lower()
    rol = (d.get('rol') or '').strip()
    grupo = (d.get('grupo') or '').strip() or None
    carrera = (d.get('carrera') or '').strip() or None
    tutor_id = d.get('tutor_id')
    if not all([nombre, apellidos, email, rol]):
        return jsonify(success=False, message='Campos requeridos'), 400
    anio = datetime.datetime.now().year
    pwd = apellidos[:4].lower() + str(anio)
    try:
        cn = db()
        if cn.execute("SELECT id FROM usuarios WHERE LOWER(email)=?", (email,)).fetchone():
            cn.close()
            return jsonify(success=False, message='Ya existe ese correo'), 409
        cn.execute("INSERT INTO usuarios (nombre,apellidos,email,password,rol,departamento_o_carrera,grupo_asignado,tutor_id) VALUES (?,?,?,?,?,?,?,?)",
                   (nombre, apellidos, email, pwd, rol, carrera, grupo, tutor_id))
        cn.commit(); cn.close()
        return jsonify(success=True, message='Usuario creado. Password: '+pwd), 201
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/usuarios/<int:uid>', methods=['DELETE'])
def delete_user(uid):
    try:
        cn = db()
        u = cn.execute("SELECT rol FROM usuarios WHERE id=?", (uid,)).fetchone()
        if not u: cn.close(); return jsonify(success=False, message='No encontrado'), 404
        if u['rol'] == 'admin': cn.close(); return jsonify(success=False, message='No se puede eliminar admin'), 403
        cn.execute("DELETE FROM usuarios WHERE id=?", (uid,))
        cn.commit(); cn.close()
        return jsonify(success=True, message='Eliminado')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/usuarios/<int:uid>', methods=['PUT'])
def update_user(uid):
    d = request.get_json() or {}
    nombre = (d.get('nombre') or '').strip()
    apellidos = (d.get('apellidos') or '').strip()
    email = (d.get('email') or '').strip().lower()
    if not all([nombre, apellidos, email]):
        return jsonify(success=False, message='Faltan campos requeridos'), 400
    try:
        cn = db()
        existe = cn.execute("SELECT id FROM usuarios WHERE LOWER(email)=? AND id!=?", (email, uid)).fetchone()
        if existe:
            cn.close()
            return jsonify(success=False, message='El correo ya está en uso por otro docente'), 409
        
        cn.execute("UPDATE usuarios SET nombre=?, apellidos=?, email=? WHERE id=?", (nombre, apellidos, email, uid))
        cn.commit()
        cn.close()
        return jsonify(success=True, message='Perfil actualizado exitosamente')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ DOCENTES ============
@app.route('/api/docentes', methods=['GET'])
def list_docentes():
    carrera = request.args.get('carrera')
    try:
        cn = db()
        q = "SELECT id,nombre,apellidos,email,departamento_o_carrera FROM usuarios WHERE rol='docente-tutor'"
        p = []
        if carrera: q += " AND departamento_o_carrera=?"; p.append(carrera)
        q += " ORDER BY nombre"
        rows = cn.execute(q, p).fetchall()
        cn.close()
        return jsonify(success=True, docentes=[dict(r) for r in rows])
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/docentes', methods=['POST'])
def add_docente():
    d = request.get_json() or {}
    nombre = (d.get('nombre') or '').strip()
    apellidos = (d.get('apellidos') or '').strip()
    email = (d.get('email') or '').strip().lower()
    depto = (d.get('departamento') or '').strip()
    if not all([nombre, apellidos, email, depto]):
        return jsonify(success=False, message='Campos requeridos'), 400
    pwd = apellidos[:4].lower() + str(datetime.datetime.now().year)
    try:
        cn = db()
        if cn.execute("SELECT id FROM usuarios WHERE LOWER(email)=?", (email,)).fetchone():
            cn.close(); return jsonify(success=False, message='Ya existe'), 409
        cn.execute("INSERT INTO usuarios (nombre,apellidos,email,password,rol,departamento_o_carrera) VALUES (?,?,?,?,'docente-tutor',?)",
                   (nombre, apellidos, email, pwd, depto))
        cn.commit(); cn.close()
        return jsonify(success=True, message='Docente registrado. Password: '+pwd), 201
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ ASIGNACIONES ============
@app.route('/api/asignaciones', methods=['GET'])
def list_asig():
    carrera = request.args.get('carrera')
    carrera_id = request.args.get('carrera_id')
    tutor_id = request.args.get('tutor_id')
    try:
        cn = db()
        q = """SELECT a.id, a.grupo, a.cuatrimestre, a.jefe_grupo_id,
                      p.nombre AS periodo, u.id AS tutor_id,
                      u.nombre||' '||u.apellidos AS tutor,
                      c.nombre AS carrera, c.id AS carrera_id,
                      jg.nombre||' '||jg.apellidos AS jefe_nombre
               FROM asignaciones a
               JOIN periodos p ON a.periodo_id=p.id
               JOIN usuarios u ON a.tutor_id=u.id
               JOIN carreras c ON a.carrera_id=c.id
               LEFT JOIN usuarios jg ON a.jefe_grupo_id=jg.id"""
        conds, params = [], []
        if carrera: conds.append("c.nombre=?"); params.append(carrera)
        if carrera_id: conds.append("a.carrera_id=?"); params.append(carrera_id)
        if tutor_id: conds.append("a.tutor_id=?"); params.append(tutor_id)
        if conds: q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY a.cuatrimestre, a.grupo"
        rows = cn.execute(q, params).fetchall()
        cn.close()
        return jsonify(success=True, asignaciones=[dict(r) for r in rows])
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/asignaciones', methods=['POST'])
def create_asig():
    """Director de Carrera asigna el grupo. No genera el PAT automáticamente, el tutor lo carga después."""
    d = request.get_json() or {}
    carrera_id = d.get('carrera_id')
    tutor_id = d.get('tutor_id')
    grupo = (d.get('grupo') or '').strip().upper()
    cuatrimestre = d.get('cuatrimestre')
    
    if not all([carrera_id, tutor_id, grupo, cuatrimestre]):
        return jsonify(success=False, message='Campos requeridos'), 400
    try:
        cn = db()
        per = cn.execute("SELECT id FROM periodos WHERE activo=1 LIMIT 1").fetchone()
        if not per: 
            cn.close()
            return jsonify(success=False, message='Sin periodo activo'), 400
        pid = per['id']

        if cn.execute("SELECT id FROM asignaciones WHERE tutor_id=? AND grupo=? AND periodo_id=?", (tutor_id, grupo, pid)).fetchone():
            cn.close()
            return jsonify(success=False, message='Ya existe esa asignacion'), 409
        
        # Simplemente insertar la asignación
        cn.execute("INSERT INTO asignaciones (carrera_id,tutor_id,grupo,periodo_id,cuatrimestre) VALUES (?,?,?,?,?)",
                   (carrera_id, tutor_id, grupo, pid, cuatrimestre))
        
        cn.commit()
        cn.close()
        return jsonify(success=True, message=f'Tutor asignado al grupo {grupo}.'), 201
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/asignaciones/<int:aid>', methods=['PUT'])
def reassign_tutor(aid):
    d = request.get_json() or {}
    nuevo_tutor_id = d.get('tutor_id')
    if not nuevo_tutor_id:
        return jsonify(success=False, message='Debe seleccionar un tutor'), 400
    try:
        cn = db()
        asig = cn.execute("SELECT * FROM asignaciones WHERE id=?", (aid,)).fetchone()
        if not asig:
            cn.close()
            return jsonify(success=False, message='Asignación no encontrada'), 404
        
        tutor = cn.execute("SELECT nombre, apellidos FROM usuarios WHERE id=?", (nuevo_tutor_id,)).fetchone()
        nombre_tutor_completo = f"{tutor['nombre']} {tutor['apellidos']}" if tutor else ""

        cn.execute("UPDATE asignaciones SET tutor_id=? WHERE id=?", (nuevo_tutor_id, aid))
        cn.execute("UPDATE pats SET nombre_tutor=? WHERE asignacion_id=?", (nombre_tutor_completo, aid))
        
        cn.commit()
        cn.close()
        return jsonify(success=True, message='Tutor reasignado correctamente')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ PLANTILLAS PAT ============
@app.route('/api/plantillas', methods=['GET'])
def list_plantillas():
    cuat = request.args.get('cuatrimestre', '1')
    pid = request.args.get('periodo_id')
    try:
        cn = db()
        q = "SELECT * FROM plantillas_pat WHERE cuatrimestre=?"
        p = [cuat]
        if pid: q += " AND periodo_id=?"; p.append(pid)
        q += " ORDER BY num_sesion"
        rows = cn.execute(q, p).fetchall()
        cn.close()
        return jsonify(success=True, plantillas=[dict(r) for r in rows])
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/plantillas', methods=['POST'])
def add_plantilla():
    d = request.get_json() or {}
    try:
        cn = db()
        per = cn.execute("SELECT id FROM periodos WHERE activo=1 LIMIT 1").fetchone()
        pid = per['id'] if per else None
        cn.execute("INSERT INTO plantillas_pat (cuatrimestre,num_sesion,corte_parcial,tematica,objetivo,resultados_esperados,periodo_id) VALUES (?,?,?,?,?,?,?)",
                   (d.get('cuatrimestre'), d.get('num_sesion'), d.get('corte_parcial',''), d.get('tematica',''), d.get('objetivo',''), d.get('resultados_esperados',''), pid))
        cn.commit(); cn.close()
        return jsonify(success=True, message='Sesion agregada'), 201
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/plantillas/<int:pid>', methods=['PUT'])
def update_plantilla(pid):
    d = request.get_json() or {}
    try:
        cn = db()
        cn.execute("UPDATE plantillas_pat SET corte_parcial=?,tematica=?,objetivo=?,resultados_esperados=? WHERE id=?",
                   (d.get('corte_parcial'), d.get('tematica'), d.get('objetivo'), d.get('resultados_esperados'), pid))
        cn.commit(); cn.close()
        return jsonify(success=True, message='Actualizado')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/plantillas/<int:pid>', methods=['DELETE'])
def del_plantilla(pid):
    try:
        cn = db()
        cn.execute("DELETE FROM plantillas_pat WHERE id=?", (pid,))
        cn.commit(); cn.close()
        return jsonify(success=True, message='Eliminado')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/plantillas/publicar', methods=['POST'])
def publicar_plantilla():
    """Director de Tutorías publica la plantilla. (No genera PATs, solo los hace disponibles)"""
    d = request.get_json() or {}
    cuat = d.get('cuatrimestre')
    try:
        cn = db()
        per = cn.execute("SELECT id FROM periodos WHERE activo=1 LIMIT 1").fetchone()
        if not per: 
            cn.close()
            return jsonify(success=False, message='Sin periodo'), 400
        
        pid = per['id']
        cn.execute("UPDATE plantillas_pat SET estado='Publicada' WHERE cuatrimestre=? AND periodo_id=?", (cuat, pid))
        
        cn.commit()
        cn.close()
        return jsonify(success=True, message=f'Plantilla publicada. (Los tutores ya pueden cargarla en su portal)')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/plantillas/limpiar', methods=['POST'])
def limpiar_borrador():
    d = request.get_json() or {}
    cuat = d.get('cuatrimestre')
    if not cuat:
        return jsonify(success=False, message='Cuatrimestre requerido'), 400
    try:
        cn = db()
        per = cn.execute("SELECT id FROM periodos WHERE activo=1 LIMIT 1").fetchone()
        if not per: cn.close(); return jsonify(success=False, message='Sin periodo'), 400
        cn.execute("DELETE FROM plantillas_pat WHERE cuatrimestre=? AND periodo_id=?", (cuat, per['id']))
        count = cn.execute("SELECT changes()").fetchone()[0]
        cn.commit(); cn.close()
        return jsonify(success=True, message='Borrador limpiado (%d sesiones eliminadas)' % count)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# --- SUBIR EXCEL PARA CREAR PLANTILLA ---
@app.route('/api/plantillas/upload-excel', methods=['POST'])
def upload_excel():
    if 'file' not in request.files:
        return jsonify(success=False, message='No se envio archivo'), 400
    f = request.files['file']
    cuat = request.form.get('cuatrimestre', '1')
    if not f.filename.endswith(('.xlsx', '.xls')):
        return jsonify(success=False, message='Solo archivos .xlsx'), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(f, data_only=True)
        ws = wb.active
        cn = db()
        per = cn.execute("SELECT id FROM periodos WHERE activo=1 LIMIT 1").fetchone()
        pid = per['id'] if per else None
        cn.execute("DELETE FROM plantillas_pat WHERE cuatrimestre=? AND periodo_id=?", (cuat, pid))
        count = 0
        for row in range(12, ws.max_row + 1):
            num = ws.cell(row, 1).value
            if num is None or not str(num).isdigit():
                continue
            corte = ws.cell(row, 3).value or ''
            tematica = ws.cell(row, 4).value or ''
            objetivo = ws.cell(row, 5).value or ''
            resultados = ws.cell(row, 6).value or ''
            if not tematica.strip():
                continue
            cn.execute("INSERT INTO plantillas_pat (cuatrimestre,num_sesion,corte_parcial,tematica,objetivo,resultados_esperados,estado,periodo_id) VALUES (?,?,?,?,?,?,'Borrador',?)",
                       (int(cuat), int(num), str(corte).strip(), str(tematica).strip(), str(objetivo).strip(), str(resultados).strip(), pid))
            count += 1
        cn.commit(); cn.close()
        return jsonify(success=True, message='%d sesiones importadas del Excel' % count)
    except ImportError:
        return jsonify(success=False, message='openpyxl no instalado. Ejecuta: pip install openpyxl'), 500
    except Exception as e:
        return jsonify(success=False, message='Error procesando Excel: ' + str(e)), 500

# --- DESCARGAR FORMATO EXCEL ---
@app.route('/api/plantillas/download-excel', methods=['GET'])
def download_excel_format():
    try:
        import io
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from flask import send_file, jsonify

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Formato PAT"

        font_title = Font(name='Arial', size=12, bold=True)
        font_bold = Font(name='Arial', size=10, bold=True)
        font_header = Font(name='Arial', size=10, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color="69B22E", end_color="69B22E", fill_type="solid") 
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        align_left_center = Alignment(horizontal='left', vertical='center', wrap_text=True)
        align_top_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        bottom_border = Border(bottom=Side(style='thin'))

        ws.merge_cells('A1:F1')
        ws['A1'] = "UNIVERSIDAD POLITÉCNICA DE FRANCISCO I. MADERO"
        ws['A1'].font = font_title
        ws['A1'].alignment = align_center

        ws.merge_cells('A2:F2')
        ws['A2'] = "DEPARTAMENTO DE ASESORÍAS Y TUTORÍAS"
        ws['A2'].font = font_title
        ws['A2'].alignment = align_center

        ws.merge_cells('A3:F3')
        ws['A3'] = "PLAN DE ACCIÓN TUTORIAL (PAT)"
        ws['A3'].font = font_title
        ws['A3'].alignment = align_center

        ws['A5'] = "PROGRAMA EDUCATIVO:"
        ws['A5'].font = font_bold
        ws.merge_cells('B5:D5')
        ws['B5'].border = bottom_border
        ws['E5'] = "CUATRIMESTRE:"
        ws['E5'].font = font_bold
        ws['F5'].border = bottom_border

        ws['A6'] = "NOMBRE DEL TUTOR:"
        ws['A6'].font = font_bold
        ws.merge_cells('B6:D6')
        ws['B6'].border = bottom_border
        ws['E6'] = "GRUPO:"
        ws['E6'].font = font_bold
        ws['F6'].border = bottom_border

        ws['A7'] = "PERIODO:"
        ws['A7'].font = font_bold
        ws.merge_cells('B7:D7')
        ws['B7'].border = bottom_border
        ws['E7'] = "FECHA:"
        ws['E7'].font = font_bold
        ws['F7'].border = bottom_border

        ws.merge_cells('A10:F10')
        ws['A10'] = "PROGRAMA CUATRIMESTRAL (CALENDARIZACIÓN DE SESIONES)"
        ws['A10'].font = font_header
        ws['A10'].fill = fill_header
        ws['A10'].alignment = align_center
        ws['A10'].border = thin_border

        headers = ["No. de sesión", "Fecha de la sesión", "Corte parcial", "Temática", "Objetivo de la sesión", "Resultados Esperados"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=11, column=col_num)
            cell.value = header
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = align_center
            cell.border = thin_border

        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 45
        ws.column_dimensions['E'].width = 45
        ws.column_dimensions['F'].width = 45
        ws.row_dimensions[11].height = 25 

        for row in range(12, 27): 
            ws.row_dimensions[row].height = 40 
            ws.cell(row=row, column=1).value = row - 11 
            ws.cell(row=row, column=1).alignment = align_center
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                if col != 1:
                    cell.alignment = align_top_left

        out = io.BytesIO()
        wb.save(out)
        out.seek(0)

        return send_file(out, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True, download_name='Formato_Plantilla_PAT.xlsx')
    except ImportError:
        from flask import jsonify
        return jsonify(success=False, message='La librería openpyxl no está instalada.'), 500
    except Exception as e:
        from flask import jsonify
        return jsonify(success=False, message='Error al generar el formato: ' + str(e)), 500

# ============ PATs ============

# NUEVO ENDPOINT: El Tutor carga la plantilla a partir de su asignación
# NUEVO ENDPOINT: El Tutor carga la plantilla a partir de su asignación
@app.route('/api/pats/cargar', methods=['POST'])
def cargar_pat_tutor():
    d = request.get_json() or {}
    asig_id = d.get('asignacion_id')
    if not asig_id: return jsonify(success=False, message='Falta asignacion_id'), 400
    
    try:
        cn = db()
        asig = cn.execute("""SELECT a.*, c.nombre as carrera_nombre, u.nombre, u.apellidos, p.nombre as periodo_nombre
                             FROM asignaciones a
                             JOIN carreras c ON a.carrera_id = c.id
                             JOIN usuarios u ON a.tutor_id = u.id
                             JOIN periodos p ON a.periodo_id = p.id
                             WHERE a.id=?""", (asig_id,)).fetchone()
        
        if not asig:
            cn.close(); return jsonify(success=False, message='Asignación no encontrada'), 404

        # 1. Verificar si el Director de Tutorías ya publicó la plantilla para este cuatrimestre
        plantillas = cn.execute("SELECT * FROM plantillas_pat WHERE cuatrimestre=? AND estado='Publicada' AND periodo_id=? ORDER BY num_sesion",
                                (asig['cuatrimestre'], asig['periodo_id'])).fetchall()

        if not plantillas:
            cn.close()
            return jsonify(success=False, message='⚠️ El Director de Tutorías aún no ha publicado la plantilla oficial para tu cuatrimestre. Intenta más tarde.'), 400

        # 2. Verificar si ya existe un PAT fantasma o vacío para esta asignación
        pat = cn.execute("SELECT id FROM pats WHERE asignacion_id=?", (asig_id,)).fetchone()
        pat_id = None
        
        if pat:
            pat_id = pat['id']
            # Si existe, verificamos si ya tiene sesiones
            ses_count = cn.execute("SELECT COUNT(*) FROM sesiones WHERE pat_id=?", (pat_id,)).fetchone()[0]
            if ses_count > 0:
                cn.close(); return jsonify(success=False, message='El PAT ya tiene sesiones cargadas previamente.'), 400
        else:
            # Si no existe, lo creamos
            tutor_nombre = f"{asig['nombre']} {asig['apellidos']}"
            cn.execute("""INSERT INTO pats (asignacion_id,cuatrimestre,programa_educativo,nombre_tutor,periodo_cuatrimestral,grupo,estado)
                          VALUES (?,?,?,?,?,?,'Borrador')""",
                       (asig_id, asig['cuatrimestre'], asig['carrera_nombre'], tutor_nombre, asig['periodo_nombre'], asig['grupo']))
            pat_id = cn.execute("SELECT last_insert_rowid()").fetchone()[0]

        # 3. Insertar todas las sesiones clonadas de la plantilla
        for pl in plantillas:
            cn.execute("INSERT INTO sesiones (pat_id,num_sesion,corte_parcial,tematica,objetivo,resultados_esperados,estado) VALUES (?,?,?,?,?,?,'Pendiente')",
                       (pat_id, pl['num_sesion'], pl['corte_parcial'], pl['tematica'], pl['objetivo'], pl['resultados_esperados']))

        cn.commit()
        cn.close()
        return jsonify(success=True, message='¡Plantilla cargada exitosamente! Ya puedes iniciar tus registros.')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500


@app.route('/api/pats', methods=['GET'])
def list_pats():
    estado = request.args.get('estado')
    carrera = request.args.get('carrera')
    tutor_id = request.args.get('tutor_id')
    try:
        cn = db()
        q = """SELECT p.id, p.cuatrimestre, p.estado, p.alumnos_h, p.alumnos_m,
                      p.programa_educativo, p.nombre_tutor, p.periodo_cuatrimestral,
                      p.grupo, p.comentarios_tutor, p.nombre_jefe_grupo,
                      a.tutor_id, c.nombre AS carrera,
                      u.nombre||' '||u.apellidos AS tutor,
                      per.nombre AS periodo
               FROM pats p
               JOIN asignaciones a ON p.asignacion_id=a.id
               JOIN carreras c ON a.carrera_id=c.id
               JOIN usuarios u ON a.tutor_id=u.id
               JOIN periodos per ON a.periodo_id=per.id"""
        conds, params = [], []
        if estado: conds.append("p.estado=?"); params.append(estado)
        if carrera: conds.append("c.nombre=?"); params.append(carrera)
        if tutor_id: conds.append("a.tutor_id=?"); params.append(tutor_id)
        if conds: q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY p.id DESC"
        rows = cn.execute(q, params).fetchall()
        cn.close()
        return jsonify(success=True, pats=[dict(r) for r in rows])
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/pats/<int:pat_id>', methods=['GET'])
def get_pat(pat_id):
    try:
        cn = db()
        p = cn.execute("""SELECT p.*, a.tutor_id, c.nombre AS carrera,
                                 u.nombre||' '||u.apellidos AS tutor_nombre_completo,
                                 per.nombre AS periodo_nombre
                          FROM pats p
                          JOIN asignaciones a ON p.asignacion_id=a.id
                          JOIN carreras c ON a.carrera_id=c.id
                          JOIN usuarios u ON a.tutor_id=u.id
                          JOIN periodos per ON a.periodo_id=per.id
                          WHERE p.id=?""", (pat_id,)).fetchone()
        if not p: cn.close(); return jsonify(success=False, message='PAT no encontrado'), 404
        sesiones = cn.execute("SELECT * FROM sesiones WHERE pat_id=? ORDER BY num_sesion", (pat_id,)).fetchall()
        cn.close()
        d = dict(p)
        d['sesiones'] = [dict(s) for s in sesiones]
        return jsonify(success=True, pat=d)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/pats', methods=['POST'])
def create_pat():
    """Fallback por si se necesita crear un PAT directamente sin botón de Carga"""
    d = request.get_json() or {}
    asig_id = d.get('asignacion_id')
    cuat = d.get('cuatrimestre')
    programa = d.get('programa_educativo', '')
    tutor_nombre = d.get('nombre_tutor', '')
    periodo = d.get('periodo_cuatrimestral', '')
    grupo = d.get('grupo', '')
    if not asig_id or not cuat:
        return jsonify(success=False, message='Asignacion y cuatrimestre requeridos'), 400
    try:
        cn = db()
        per = cn.execute("SELECT id FROM periodos WHERE activo=1 LIMIT 1").fetchone()
        pid = per['id'] if per else None
        if cn.execute("SELECT id FROM pats WHERE asignacion_id=?", (asig_id,)).fetchone():
            cn.close(); return jsonify(success=False, message='Ya existe un PAT para esta asignacion'), 409
        cn.execute("""INSERT INTO pats (asignacion_id,cuatrimestre,programa_educativo,nombre_tutor,periodo_cuatrimestral,grupo,estado)
                      VALUES (?,?,?,?,?,?,'Borrador')""",
                   (asig_id, cuat, programa, tutor_nombre, periodo, grupo))
        pat_id = cn.execute("SELECT last_insert_rowid()").fetchone()[0]
        plantillas = cn.execute("SELECT * FROM plantillas_pat WHERE cuatrimestre=? AND estado='Publicada' AND periodo_id=? ORDER BY num_sesion",
                                (cuat, pid)).fetchall()
        for pl in plantillas:
            cn.execute("INSERT INTO sesiones (pat_id,num_sesion,corte_parcial,tematica,objetivo,resultados_esperados,estado) VALUES (?,?,?,?,?,?,'Pendiente')",
                       (pat_id, pl['num_sesion'], pl['corte_parcial'], pl['tematica'], pl['objetivo'], pl['resultados_esperados']))
        cn.commit(); cn.close()
        return jsonify(success=True, message='PAT creado con %d sesiones' % len(plantillas), id=pat_id), 201
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/pats/<int:pat_id>', methods=['PATCH'])
def update_pat(pat_id):
    d = request.get_json() or {}
    campos = ['alumnos_h','alumnos_m','comentarios_tutor','estado','observaciones_director',
              'programa_educativo','nombre_tutor','periodo_cuatrimestral','grupo','nombre_jefe_grupo']
    updates = {k: v for k, v in d.items() if k in campos}
    if 'estado' in updates:
        if updates['estado'] == 'Pendiente Revision':
            updates['fecha_envio'] = datetime.datetime.now().isoformat()
        elif updates['estado'] == 'Aprobado':
            updates['fecha_aprobacion'] = datetime.datetime.now().isoformat()
    if not updates:
        return jsonify(success=False, message='Sin campos validos'), 400
    try:
        cn = db()
        sets = ', '.join('%s=?' % k for k in updates)
        vals = list(updates.values()) + [pat_id]
        cn.execute("UPDATE pats SET %s WHERE id=?" % sets, vals)
        cn.commit(); cn.close()
        return jsonify(success=True, message='PAT actualizado')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ SESIONES ============
@app.route('/api/sesiones/<int:sid>', methods=['PATCH'])
def update_sesion(sid):
    d = request.get_json() or {}
    campos = ['fecha_programada','hora_programada','resultados_obtenidos','canalizaciones',
              'estado','motivo_cancelacion','firma_jefe','fecha_validacion',
              'tematica','objetivo','resultados_esperados','corte_parcial']
    updates = {k: v for k, v in d.items() if k in campos}
    if not updates:
        return jsonify(success=False, message='Sin campos'), 400
    try:
        cn = db()
        sets = ', '.join('%s=?' % k for k in updates)
        vals = list(updates.values()) + [sid]
        r = cn.execute("UPDATE sesiones SET %s WHERE id=?" % sets, vals)
        if r.rowcount == 0: cn.close(); return jsonify(success=False, message='Sesion no encontrada'), 404
        cn.commit(); cn.close()
        return jsonify(success=True, message='Sesion actualizada')
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ SOLICITUDES JEFE DE GRUPO ============
@app.route('/api/solicitudes-jefe', methods=['GET'])
def list_sol():
    estado = request.args.get('estado')
    sol_id = request.args.get('solicitante_id')
    try:
        cn = db()
        q = "SELECT s.*, u.nombre||' '||u.apellidos AS solicitante_nombre FROM solicitudes_jefe s JOIN usuarios u ON s.solicitante_id=u.id"
        conds, params = [], []
        if estado: conds.append("s.estado=?"); params.append(estado)
        if sol_id: conds.append("s.solicitante_id=?"); params.append(sol_id)
        if conds: q += " WHERE " + " AND ".join(conds)
        q += " ORDER BY s.id DESC"
        rows = cn.execute(q, params).fetchall()
        cn.close()
        return jsonify(success=True, solicitudes=[dict(r) for r in rows])
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/solicitudes-jefe', methods=['POST'])
def create_sol():
    d = request.get_json() or {}
    try:
        cn = db()
        cn.execute("INSERT INTO solicitudes_jefe (solicitante_id,nombre_alumno,apellidos_alumno,email_alumno,grupo,carrera) VALUES (?,?,?,?,?,?)",
                   (d.get('solicitante_id'), d.get('nombre_alumno',''), d.get('apellidos_alumno',''), d.get('email_alumno',''), d.get('grupo',''), d.get('carrera','')))
        cn.commit(); cn.close()
        return jsonify(success=True, message='Solicitud enviada'), 201
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

@app.route('/api/solicitudes-jefe/<int:sid>', methods=['PATCH'])
def review_sol(sid):
    d = request.get_json() or {}
    estado = d.get('estado')
    if estado not in ('Aprobada', 'Rechazada'):
        return jsonify(success=False, message='Estado no valido'), 400
    try:
        cn = db()
        sol = cn.execute("SELECT * FROM solicitudes_jefe WHERE id=?", (sid,)).fetchone()
        if not sol: cn.close(); return jsonify(success=False, message='No encontrada'), 404
        cn.execute("UPDATE solicitudes_jefe SET estado=?, observaciones=? WHERE id=?",
                   (estado, d.get('observaciones',''), sid))
        msg = 'Solicitud ' + estado.lower()
        if estado == 'Aprobada':
            pwd = (sol['apellidos_alumno'][:4].lower() + str(datetime.datetime.now().year))
            cn.execute("INSERT INTO usuarios (nombre,apellidos,email,password,rol,departamento_o_carrera,grupo_asignado,tutor_id) VALUES (?,?,?,?,'jefe-grupo',?,?,?)",
                       (sol['nombre_alumno'], sol['apellidos_alumno'], sol['email_alumno'], pwd, sol['carrera'], sol['grupo'], sol['solicitante_id']))
            jefe_id = cn.execute("SELECT last_insert_rowid()").fetchone()[0]
            cn.execute("UPDATE asignaciones SET jefe_grupo_id=? WHERE grupo=? AND tutor_id=?",
                       (jefe_id, sol['grupo'], sol['solicitante_id']))
            msg += '. Cuenta creada con password: ' + pwd
        cn.commit(); cn.close()
        return jsonify(success=True, message=msg)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ STATS ============
@app.route('/api/stats')
def stats():
    try:
        cn = db()
        r = {
            'usuarios': cn.execute("SELECT COUNT(*) FROM usuarios WHERE activo=1").fetchone()[0],
            'pats': cn.execute("SELECT COUNT(*) FROM pats").fetchone()[0],
            'pats_aprobados': cn.execute("SELECT COUNT(*) FROM pats WHERE estado='Aprobado'").fetchone()[0],
            'solicitudes_pend': cn.execute("SELECT COUNT(*) FROM solicitudes_jefe WHERE estado='Pendiente'").fetchone()[0],
        }
        cn.close()
        return jsonify(success=True, stats=r)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

# ============ SISTEMA DE AUDITORÍAS (SNAPSHOTS) ============
try:
    _cn = sqlite3.connect(RUTA_BD)
    _cn.execute("ALTER TABLE pats ADD COLUMN snapshot_auditoria TEXT")
    _cn.commit()
    _cn.close()
except:
    pass 

@app.route('/api/pats/<int:pat_id>/enviar-auditoria', methods=['POST'])
def enviar_auditoria(pat_id):
    try:
        cn = db()
        cur_ses = cn.execute("SELECT * FROM sesiones WHERE pat_id=? ORDER BY num_sesion", (pat_id,))
        sesiones = [dict(row) for row in cur_ses.fetchall()]
        import datetime, json
        snapshot = {
            "fecha_envio": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "sesiones": sesiones
        }
        cn.execute("UPDATE pats SET snapshot_auditoria=? WHERE id=?", (json.dumps(snapshot), pat_id))
        cn.commit()
        cn.close()
        return jsonify(success=True, message="Avance enviado a Dirección para Auditoría")
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

if __name__ == '__main__':
    print("============================================================")
    print("  Servidor Sistema PAT - UPFIM")
    print("  BD: " + RUTA_BD)
    print("  http://localhost:5000")
    print("============================================================")
    app.run(debug=True, port=5000)